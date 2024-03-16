import openpyxl
import jiagu
import pandas as pd


workbook = openpyxl.load_workbook('files/comments.xlsx')

sheet = workbook.active


max_row = sheet.max_row

# 遍历每行，逐行读取第一列文本并进行情感分析，将结果存储到第二列
for row in range(2, max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value

    sentiment = jiagu.sentiment(cell_value)

    # 将情感分数存储到第二列
    sheet.cell(row=row, column=2, value=sentiment[0])


workbook.save('emotion/comments_emotion.xlsx')


workbook.close()



input_excel_path = 'emotion/comments_emotion.xlsx'
df = pd.read_excel(input_excel_path)


filtered_df = df[(df.iloc[:, 1] == 'positive')]
output_excel_path = 'emotion/emotion_positive.xlsx'
filtered_df.to_excel(output_excel_path, index=False)


filtered_df = df[(df.iloc[:, 1] == 'negative')]
output_excel_path = 'emotion/emotion_negative.xlsx'
filtered_df.to_excel(output_excel_path, index=False)


print("情感分析完成")


